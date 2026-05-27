"""
loaders.excel_loader — .xlsx, .xls, .xlsm ingestion.

Engine selection:
    .xlsx, .xlsm → openpyxl (reads values; macros preserved on write-back)
    .xls         → xlrd (only versions ≤ 2.0.1 read .xls; legacy format)

Multi-sheet workbooks are fully parsed; the loader picks the most-populated
non-empty sheet as the primary. Callers can switch via
LoadedSpreadsheet.set_active_sheet().
"""

from __future__ import annotations

import time
from pathlib import Path

import pandas as pd

from utils import get_logger

from .base_loader import (
    BaseLoader,
    CorruptedFileError,
    EmptyFileError,
    LoadedSpreadsheet,
    LoaderError,
    SheetInfo,
)
from .normalize import normalize_dataframe

log = get_logger("loaders.excel")


_XLSX_EXTS = (".xlsx", ".xlsm")
_XLS_EXTS  = (".xls",)


class ExcelLoader(BaseLoader):
    """Loader for .xlsx / .xls / .xlsm workbooks."""

    EXTENSIONS = (".xlsx", ".xls", ".xlsm")
    NAME = "ExcelLoader"

    # ------------------------------------------------------------------
    # Engine selection
    # ------------------------------------------------------------------

    def _engine(self) -> str:
        suffix = self.file_path.suffix.lower()
        if suffix in _XLSX_EXTS:
            return "openpyxl"
        if suffix in _XLS_EXTS:
            return "xlrd"
        raise LoaderError(f"Unknown Excel extension: {suffix}")

    def _file_type(self) -> str:
        return self.file_path.suffix.lower().lstrip(".")

    # ------------------------------------------------------------------
    # Required API
    # ------------------------------------------------------------------

    def extract_sheets(self) -> list[str]:
        """List sheet names without reading data (cheap)."""
        try:
            with pd.ExcelFile(self.file_path, engine=self._engine()) as xf:
                return list(xf.sheet_names)
        except ImportError as exc:
            raise LoaderError(
                f"Missing engine for {self.file_path.suffix}: {exc}. "
                f"Install with: pip install xlrd"
            ) from exc
        except Exception as exc:
            raise CorruptedFileError(
                f"Could not read workbook {self.file_path.name}: {exc}"
            ) from exc

    def load(self, sheet: str | None = None, *, max_rows: int | None = None) -> LoadedSpreadsheet:
        t0 = time.perf_counter()
        engine = self._engine()
        path   = self.file_path
        warnings: list[str] = []

        log.info("excel: loading %s via %s", path.name, engine)

        try:
            xf = pd.ExcelFile(path, engine=engine)
        except ImportError as exc:
            raise LoaderError(
                f"Missing engine '{engine}' for {path.suffix}. "
                f"Install with: pip install {'xlrd' if engine == 'xlrd' else 'openpyxl'}"
            ) from exc
        except Exception as exc:
            raise CorruptedFileError(
                f"Could not open workbook {path.name}: {exc}"
            ) from exc

        sheet_names = list(xf.sheet_names)
        if not sheet_names:
            raise EmptyFileError(f"Workbook {path.name} has no sheets")

        if sheet is not None and sheet not in sheet_names:
            raise LoaderError(
                f"Sheet {sheet!r} not found in {path.name}. Available: {sheet_names}"
            )

        # Parse every sheet so callers can switch without re-opening the file.
        sheets: dict[str, pd.DataFrame] = {}
        infos:  list[SheetInfo]         = []

        for name in sheet_names:
            try:
                df = xf.parse(name, nrows=max_rows)
            except Exception as exc:
                warnings.append(f"Sheet {name!r} could not be parsed: {exc}")
                log.warning("excel: skipping sheet %r — %s", name, exc)
                infos.append(SheetInfo(name=name, rows=0, columns=0,
                                       column_names=[], is_empty=True))
                continue

            df, w = normalize_dataframe(df)
            warnings.extend(f"[{name}] {msg}" for msg in w)

            sheets[name] = df
            infos.append(SheetInfo(
                name=name,
                rows=len(df),
                columns=len(df.columns),
                column_names=[str(c) for c in df.columns],
                is_empty=df.empty,
            ))
            log.debug("excel: sheet %r → %d rows × %d cols", name, len(df), len(df.columns))

        try:
            xf.close()
        except Exception:
            pass

        non_empty = [info for info in infos if not info.is_empty]
        if not non_empty:
            raise EmptyFileError(f"Workbook {path.name} contains no usable data on any sheet")

        # Pick primary sheet
        if sheet is not None:
            primary = sheet
        else:
            primary = max(non_empty, key=lambda i: (i.rows, i.columns)).name
        for info in infos:
            info.is_primary = (info.name == primary)

        meta = self._workbook_metadata()

        elapsed = time.perf_counter() - t0
        log.info("excel: loaded %d sheet(s), primary=%r in %.2fs",
                 len(sheets), primary, elapsed)

        return LoadedSpreadsheet(
            file_path=str(path),
            file_type=self._file_type(),
            loader_name=self.NAME,
            sheets=sheets,
            sheet_info=infos,
            active_sheet=primary,
            workbook_metadata=meta,
            elapsed=elapsed,
            warnings=warnings,
        )

    # ------------------------------------------------------------------
    # Workbook metadata
    # ------------------------------------------------------------------

    def _workbook_metadata(self) -> dict:
        """Pull title/author/dates from .xlsx/.xlsm via openpyxl. Best-effort."""
        if self.file_path.suffix.lower() not in _XLSX_EXTS:
            return {}
        try:
            import openpyxl  # noqa: PLC0415
            wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            props = wb.properties
            meta = {
                "title":       props.title or "",
                "creator":     props.creator or "",
                "last_modified_by": props.lastModifiedBy or "",
                "created":     props.created.isoformat() if props.created else "",
                "modified":    props.modified.isoformat() if props.modified else "",
                "has_macros":  bool(getattr(wb, "vba_archive", None)),
            }
            wb.close()
            return meta
        except Exception as exc:
            log.debug("excel: metadata read failed — %s", exc)
            return {}


# ---------------------------------------------------------------------------
# Write-back helpers (used by excel_writer)
# ---------------------------------------------------------------------------


def write_xlsx(df: pd.DataFrame, path: str | Path, sheet_name: str = "Sheet1") -> None:
    """Plain .xlsx write via openpyxl."""
    p = Path(path)
    df.to_excel(p, index=False, sheet_name=sheet_name, engine="openpyxl")


def write_xlsm_preserving_macros(
    df: pd.DataFrame,
    path: str | Path,
    sheet_name: str | None = None,
) -> None:
    """
    Write `df` into an existing .xlsm workbook while preserving its VBA project.

    Strategy:
      1. Open the original workbook with keep_vba=True (preserves macros).
      2. If `sheet_name` exists, clear its rows and rewrite header + data.
         Otherwise create it.
      3. Save back to the same .xlsm path (still macro-enabled).

    The original workbook must exist at `path` for macros to be preserved.
    If it does not, falls back to a plain .xlsx write.
    """
    import openpyxl  # noqa: PLC0415
    from openpyxl.utils.dataframe import dataframe_to_rows  # noqa: PLC0415

    p = Path(path)
    if p.suffix.lower() != ".xlsm":
        raise LoaderError(f"write_xlsm_preserving_macros: not an .xlsm path: {p}")

    if not p.exists():
        log.warning("excel: .xlsm target missing — falling back to xlsx write at %s", p)
        write_xlsx(df, p.with_suffix(".xlsx"))
        return

    wb = openpyxl.load_workbook(p, keep_vba=True)
    target = sheet_name or wb.sheetnames[0]

    if target in wb.sheetnames:
        ws = wb[target]
        ws.delete_rows(1, ws.max_row or 1)
    else:
        ws = wb.create_sheet(target)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    wb.save(p)
    wb.close()
    log.info("excel: wrote %d rows to .xlsm sheet %r (macros preserved)", len(df), target)
